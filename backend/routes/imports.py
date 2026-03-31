"""
backend/routes/imports.py
==========================
Excel import pipeline and template download endpoints.

POST /api/import/upload          — upload & process an Excel file
GET  /api/import/template/<type> — download a blank Excel template
GET  /api/import/logs            — import history
GET  /api/import/failed/<log_id> — failed records for a specific import
"""

import io
import json
import logging
import os
import tempfile
import uuid
from datetime import datetime
from typing import Optional

import pandas as pd
from flask import Blueprint, request, jsonify, session, send_file
from werkzeug.utils import secure_filename

from backend.models.database import get_connection
from backend.services.helpers import (
    require_auth, sanitize_str, rows_to_list, make_reference
)

logger = logging.getLogger(__name__)
imports_bp = Blueprint("imports", __name__)

ALLOWED_EXTENSIONS = {".xlsx", ".xls"}

# ── Column Definitions ────────────────────────────────────────────────────────

MANDATORY_COLUMNS = {
    "sales": ["date", "product_name", "quantity", "price"],
    "inventory": ["sku", "quantity"],
    "products": ["name", "sell_price"],
}

OPTIONAL_COLUMNS = {
    "sales": ["customer_name", "payment_method", "salesperson", "status"],
    "inventory": ["location"],
    "products": ["category", "cost_price", "reorder_point", "sku"],
}


def _allowed_file(filename: str) -> bool:
    _, ext = os.path.splitext(filename.lower())
    return ext in ALLOWED_EXTENSIONS


# ── Upload & Process ──────────────────────────────────────────────────────────

@imports_bp.route("/upload", methods=["POST"])
@require_auth(roles=["admin", "staff"])
def upload_excel():
    """
    POST /api/import/upload  (multipart/form-data)
    Fields:
      - file        : Excel file
      - import_type : "sales" | "inventory" | "products"
    """
    import_type = sanitize_str(request.form.get("import_type", ""), max_len=20)
    if import_type not in ("sales", "inventory", "products"):
        return jsonify({"success": False, "error": "import_type must be sales, inventory, or products."}), 400

    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded."}), 400

    file = request.files["file"]
    if not file.filename or not _allowed_file(file.filename):
        return jsonify({"success": False, "error": "Only .xlsx and .xls files are accepted."}), 400

    # Save file to temp location
    safe_name = secure_filename(file.filename)
    tmp_path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4().hex}_{safe_name}")
    file.save(tmp_path)

    try:
        result = _process_excel(tmp_path, import_type)
    except Exception as exc:
        logger.exception("Unexpected error during Excel import")
        return jsonify({"success": False, "error": str(exc)}), 500
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    return jsonify(result)


def _process_excel(file_path: str, import_type: str) -> dict:
    """
    Core import pipeline:
      1. Load Excel into DataFrame
      2. Validate mandatory columns
      3. Iterate rows: validate, transform, insert / record failure
      4. Log summary to excel_import_logs
    """
    try:
        df = pd.read_excel(file_path, engine="openpyxl")
    except Exception as e:
        raise ValueError(f"Cannot read Excel file: {e}") from e

    # Normalise column names — strip trailing " *" added by the template for required columns
    df.columns = [str(c).strip().rstrip('*').strip().lower().replace(" ", "_") for c in df.columns]
    df = df.dropna(how="all")  # Drop completely empty rows

    # Drop the notes/hints row injected by the template as the first data row.
    # Check only the very first row: if any mandatory column contains a string
    # longer than 25 characters it is a description/hint, not real data.
    if len(df) > 0:
        mandatory_present = [c for c in MANDATORY_COLUMNS.get(import_type, []) if c in df.columns]
        first = df.iloc[0]
        if any(isinstance(first.get(c), str) and len(str(first.get(c, ""))) > 25
               for c in mandatory_present):
            df = df.iloc[1:].reset_index(drop=True)

    mandatory = MANDATORY_COLUMNS[import_type]
    missing = [col for col in mandatory if col not in df.columns]
    if missing:
        raise ValueError(
            f"Missing mandatory columns: {', '.join(missing)}. "
            f"Required: {', '.join(mandatory)}"
        )

    total_rows = len(df)
    success_count = 0
    failed_records: list = []

    with get_connection() as conn:
        for idx, row in df.iterrows():
            row_num = int(idx) + 2  # Excel row 1 = header, so data starts at 2
            try:
                if import_type == "sales":
                    _import_sales_row(conn, row, row_num)
                elif import_type == "inventory":
                    _import_inventory_row(conn, row, row_num)
                elif import_type == "products":
                    _import_products_row(conn, row, row_num)
                success_count += 1
            except _ImportRowError as e:
                failed_records.append({
                    "row_number": row_num,
                    "failure_reason": str(e),
                    "raw_data": json.dumps({k: str(v) for k, v in row.items()}),
                })
            except Exception as e:
                failed_records.append({
                    "row_number": row_num,
                    "failure_reason": f"Unexpected error: {e}",
                    "raw_data": json.dumps({k: str(v) for k, v in row.items()}),
                })

        failed_count = len(failed_records)

        # Write import log
        cursor = conn.execute(
            """
            INSERT INTO excel_import_logs
              (filename, import_type, total_rows, success_rows, failed_rows,
               status, imported_by)
            VALUES (?, ?, ?, ?, ?, 'completed', ?)
            """,
            (
                os.path.basename(file_path),
                import_type,
                total_rows,
                success_count,
                failed_count,
                session.get("user_id"),
            ),
        )
        log_id = cursor.lastrowid

        # Write failed records
        if failed_records:
            conn.executemany(
                """
                INSERT INTO failed_import_records (import_log_id, row_number, failure_reason, raw_data)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (log_id, fr["row_number"], fr["failure_reason"], fr["raw_data"])
                    for fr in failed_records
                ],
            )

    logger.info(
        "Import complete: type=%s total=%d success=%d failed=%d",
        import_type, total_rows, success_count, failed_count,
    )

    return {
        "success": True,
        "log_id": log_id,
        "import_type": import_type,
        "total_rows": total_rows,
        "success_rows": success_count,
        "failed_rows": failed_count,
        "failed_records": failed_records,
    }


class _ImportRowError(Exception):
    """Raised for expected, row-level validation failures."""


def _clean_str(val, max_len: int = 300) -> str:
    if pd.isna(val):
        return ""
    return str(val).strip()[:max_len]


def _clean_float(val) -> float:
    if pd.isna(val):
        raise _ImportRowError("Numeric field is empty.")
    try:
        return float(val)
    except (ValueError, TypeError):
        raise _ImportRowError(f"Invalid number: {val}")


def _clean_int(val) -> int:
    return int(_clean_float(val))


def _import_sales_row(conn, row, row_num: int) -> None:
    """Insert one sales row; raise _ImportRowError on any validation failure."""
    raw_date = _clean_str(row.get("date", ""))
    product_name = _clean_str(row.get("product_name", ""))
    quantity = _clean_int(row.get("quantity", 0))
    price = _clean_float(row.get("price", 0))

    if not raw_date:
        raise _ImportRowError("date is empty.")
    if not product_name:
        raise _ImportRowError("product_name is empty.")
    if quantity <= 0:
        raise _ImportRowError("quantity must be > 0.")
    if price <= 0:
        raise _ImportRowError("price must be > 0.")

    # Normalise date
    try:
        sale_date = pd.to_datetime(raw_date).strftime("%Y-%m-%d")
    except Exception:
        raise _ImportRowError(f"Cannot parse date: {raw_date}")

    # Look up product
    product = conn.execute(
        "SELECT id FROM products WHERE name=? AND is_active=1", (product_name,)
    ).fetchone()
    if not product:
        raise _ImportRowError(f"Product '{product_name}' not found in catalogue.")

    reference = make_reference("TRD")
    # Check duplicate by (product_id, sale_date, quantity, price) as a reasonable dedup key
    dup = conn.execute(
        """
        SELECT id FROM sales
        WHERE product_id=? AND sale_date=? AND quantity=? AND unit_price=?
        """,
        (product["id"], sale_date, quantity, price),
    ).fetchone()
    if dup:
        raise _ImportRowError("Duplicate record — already exists in database.")

    customer_name = _clean_str(row.get("customer_name", ""), max_len=200)
    payment_method = _clean_str(row.get("payment_method", "Cash"), max_len=50) or "Cash"
    salesperson = _clean_str(row.get("salesperson", ""), max_len=200)

    # Honour optional status column; default to 'completed'
    raw_status = _clean_str(row.get("status", "completed"), max_len=20).lower()
    if raw_status not in ("completed", "pending", "cancelled"):
        raw_status = "completed"

    # For completed sales: check available stock BEFORE inserting
    if raw_status == "completed":
        inv = conn.execute(
            "SELECT COALESCE(quantity, 0) AS qty FROM inventory WHERE product_id=?",
            (product["id"],),
        ).fetchone()
        available = inv["qty"] if inv else 0
        if available < quantity:
            raise _ImportRowError(
                f"Insufficient stock for '{product_name}'. "
                f"Available: {available}, requested: {quantity}. "
                f"Adjust stock first or import as 'pending'."
            )

    cursor = conn.execute(
        """
        INSERT INTO sales
          (reference, product_id, quantity, unit_price, customer_name,
           payment_method, salesperson, status, sale_date, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (reference, product["id"], quantity, price, customer_name,
         payment_method, salesperson, raw_status, sale_date, session.get("user_id")),
    )
    sale_id = cursor.lastrowid

    # Only deduct stock and record movement for completed sales
    if raw_status == "completed":
        conn.execute(
            """
            UPDATE inventory SET quantity = quantity - ?, updated_at = datetime('now')
            WHERE product_id = ?
            """,
            (quantity, product["id"]),
        )
        conn.execute(
            """
            INSERT INTO inventory_movements
              (product_id, movement_type, quantity_delta, reference_id, note, moved_by)
            VALUES (?, 'sale', ?, ?, 'Excel import', ?)
            """,
            (product["id"], -quantity, sale_id, session.get("user_id")),
        )


def _import_inventory_row(conn, row, row_num: int) -> None:
    sku = _clean_str(row.get("sku", ""))
    quantity = _clean_int(row.get("quantity", 0))

    if not sku:
        raise _ImportRowError("sku is empty.")

    product = conn.execute("SELECT id FROM products WHERE sku=?", (sku,)).fetchone()
    if not product:
        raise _ImportRowError(f"SKU '{sku}' not found.")

    location = _clean_str(row.get("location", "Main Warehouse"), max_len=200) or "Main Warehouse"

    conn.execute(
        """
        INSERT INTO inventory (product_id, quantity, location)
        VALUES (?, ?, ?)
        ON CONFLICT(product_id) DO UPDATE
        SET quantity=excluded.quantity, location=excluded.location, updated_at=datetime('now')
        """,
        (product["id"], quantity, location),
    )


def _import_products_row(conn, row, row_num: int) -> None:
    import uuid as _uuid
    name = _clean_str(row.get("name", ""))
    sell_price = _clean_float(row.get("sell_price", 0))

    if not name:
        raise _ImportRowError("name is empty.")
    if sell_price <= 0:
        raise _ImportRowError("sell_price must be > 0.")

    sku = _clean_str(row.get("sku", ""), max_len=50) or ("SKU-" + _uuid.uuid4().hex[:6].upper())
    category = _clean_str(row.get("category", "General"), max_len=100) or "General"
    cost_price = float(row.get("cost_price", 0) or 0)
    reorder_point = int(row.get("reorder_point", 10) or 10)

    dup = conn.execute("SELECT id FROM products WHERE sku=?", (sku,)).fetchone()
    if dup:
        raise _ImportRowError(f"SKU '{sku}' already exists.")

    cursor = conn.execute(
        """
        INSERT INTO products (sku, name, category, sell_price, cost_price, reorder_point)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (sku, name, category, sell_price, cost_price, reorder_point),
    )
    conn.execute(
        "INSERT INTO inventory (product_id, quantity) VALUES (?, 0)", (cursor.lastrowid,)
    )


# ── Template Download ─────────────────────────────────────────────────────────

@imports_bp.route("/template/<import_type>", methods=["GET"])
@require_auth()
def download_template(import_type: str):
    """GET /api/import/template/sales|inventory|products — returns xlsx template.

    Produces a formatted workbook with:
      - Row 1 : column names  (green  = required, yellow = optional)
      - Row 2 : notes/hints   (what each column expects)
      - Rows 3+: sample data  (two realistic rows)
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    if import_type not in ("sales", "inventory", "products"):
        return jsonify({"error": "Invalid template type."}), 400

    # ── Schema definitions ────────────────────────────────────────────────────
    # Each entry: (column_name, required, note, sample_value_row1, sample_value_row2)
    SCHEMAS = {
        "sales": [
            ("date",           True,  "Sale date — YYYY-MM-DD format (e.g. 2026-03-28)",               "2026-03-01",   "2026-03-15"),
            ("product_name",   True,  "Exact product name as it appears in the catalogue",              "Laptop Pro 15","Wireless Mouse"),
            ("quantity",       True,  "Number of units sold — must be a whole number > 0",              10,             3),
            ("price",          True,  "Unit selling price in KES — must be > 0",                        85000.00,       1500.00),
            ("customer_name",  False, "Customer full name or company name (optional)",                  "Acme Corp",    "Jane Wanjiku"),
            ("payment_method", False, "Cash | M-Pesa | Bank Transfer | Credit | Other (default: Cash)","M-Pesa",       "Cash"),
            ("salesperson",    False, "Name of the staff member who made the sale (optional)",          "Alice Muthoni","Bob Otieno"),
            ("status",         False, "completed | pending | cancelled (default: completed)",           "completed",    "completed"),
        ],
        "inventory": [
            ("sku",      True,  "Product SKU — must exactly match an existing product SKU",    "SKU-001",    "SKU-002"),
            ("quantity", True,  "Stock quantity to set — whole number, 0 or more",             250,          80),
            ("location", False, "Storage location name (default: Main Warehouse)",             "Main Warehouse", "Warehouse B"),
        ],
        "products": [
            ("name",          True,  "Full product name — must be unique",                                 "Laptop Pro 15",  "Wireless Mouse"),
            ("sell_price",    True,  "Retail selling price in KES — must be > 0",                         85000.00,         1500.00),
            ("cost_price",    False, "Purchase / landed cost in KES — used for profit calculations",       62000.00,         800.00),
            ("category",      False, "Product category (default: General)",                               "Electronics",    "Accessories"),
            ("reorder_point", False, "Minimum stock level before a low-stock alert is raised (default: 10)", 5,             20),
            ("sku",           False, "Unique SKU / product code — auto-generated if left blank",           "SKU-LAP-001",    "SKU-MOU-001"),
        ],
    }

    schema = SCHEMAS[import_type]
    columns     = [s[0] for s in schema]
    required    = [s[1] for s in schema]
    notes       = [s[2] for s in schema]
    sample_row1 = [s[3] for s in schema]
    sample_row2 = [s[4] for s in schema]

    # ── Styles ────────────────────────────────────────────────────────────────
    REQUIRED_FILL = PatternFill("solid", fgColor="1A7C3E")   # dark green
    OPTIONAL_FILL = PatternFill("solid", fgColor="B8860B")   # dark goldenrod
    NOTE_FILL     = PatternFill("solid", fgColor="F0F4F0")   # very light green
    HDR_FONT      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    NOTE_FONT     = Font(name="Calibri", italic=True, color="555555", size=9)
    DATA_FONT     = Font(name="Calibri", size=10)
    CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin          = Side(style="thin", color="CCCCCC")
    BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()

    # ── Data sheet ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Data"
    ws.freeze_panes = "A3"   # freeze header + notes rows

    # Row 1 — column headers
    for col_idx, (col_name, is_req) in enumerate(zip(columns, required), start=1):
        cell = ws.cell(row=1, column=col_idx,
                       value=col_name + (" *" if is_req else ""))
        cell.fill      = REQUIRED_FILL if is_req else OPTIONAL_FILL
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        cell.border    = BORDER

    # Row 2 — notes
    for col_idx, note in enumerate(notes, start=1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.fill      = NOTE_FILL
        cell.font      = NOTE_FONT
        cell.alignment = LEFT
        cell.border    = BORDER

    # Rows 3-4 — sample data
    for row_idx, row_data in enumerate([sample_row1, sample_row2], start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DATA_FONT
            cell.alignment = LEFT
            cell.border    = BORDER

    # Column widths — fit to the longest of: header, note (capped), sample
    for col_idx, (col_name, note, s1) in enumerate(zip(columns, notes, sample_row1), start=1):
        max_w = max(len(col_name) + 4, min(len(note), 48), len(str(s1)) + 4)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 40   # notes row — taller for wrapped text

    # ── Legend sheet ─────────────────────────────────────────────────────────
    wl = wb.create_sheet("Legend")
    wl.column_dimensions["A"].width = 22
    wl.column_dimensions["B"].width = 60

    legend_rows = [
        ("LEGEND", ""),
        ("Green header ( * )", "Required column — must be filled in for every row"),
        ("Yellow header",      "Optional column — leave blank to use the default value shown in the note"),
        ("Row 2",              "Notes row — describes expected format and valid values for each column"),
        ("Rows 3+",            "Your data — delete sample rows and paste or type your records here"),
        ("", ""),
        ("TIPS", ""),
        ("Do not rename columns", "The column names in row 1 must stay exactly as provided"),
        ("Do not delete row 1",   "The header row is required; the notes row (row 2) can be deleted"),
        ("Dates",                  "Use YYYY-MM-DD format, e.g. 2026-03-28"),
        ("Numbers",                "Do not include currency symbols or commas in price/quantity fields"),
        ("Blank optional cells",   "Leave blank (do not type 'N/A') — defaults will be applied automatically"),
    ]

    bold_rows = {0, 6}
    for r_idx, (label, desc) in enumerate(legend_rows, start=1):
        a = wl.cell(row=r_idx, column=1, value=label)
        b = wl.cell(row=r_idx, column=2, value=desc)
        if r_idx - 1 in bold_rows:
            a.font = Font(bold=True, size=11)
        else:
            a.font = Font(size=10)
            b.font = Font(size=10)
        a.alignment = LEFT
        b.alignment = LEFT

    # Colour the legend swatch cells
    wl["A2"].fill = REQUIRED_FILL
    wl["A2"].font = Font(bold=True, color="FFFFFF", size=10)
    wl["A3"].fill = OPTIONAL_FILL
    wl["A3"].font = Font(bold=True, color="FFFFFF", size=10)

    # Return to Data sheet as active
    wb.active = ws

    # ── Serialise & send ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name=f"tradedesk_{import_type}_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Import Logs ───────────────────────────────────────────────────────────────

@imports_bp.route("/logs", methods=["GET"])
@require_auth()
def list_import_logs():
    with get_connection() as conn:
        rows = conn.execute("""
            SELECT l.*, u.username AS imported_by_name
            FROM excel_import_logs l LEFT JOIN users u ON l.imported_by = u.id
            ORDER BY l.imported_at DESC LIMIT 50
        """).fetchall()
    return jsonify(rows_to_list(rows))


@imports_bp.route("/failed/<int:log_id>", methods=["GET"])
@require_auth()
def list_failed_records(log_id: int):
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM failed_import_records WHERE import_log_id=? ORDER BY row_number",
            (log_id,),
        ).fetchall()
    return jsonify(rows_to_list(rows))
